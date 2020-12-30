import openpyxl
import sys
from argparse import Namespace

def generate_args():
    args = Namespace(label1="P",
                     dataset1="../dataset/PTJ_prep.full-gen.test.best.xlsx",
                     label2="S",
                     dataset2="../dataset/SMK_prep.full-gen.test.best.xlsx"
                     )

def make_problem_for_probID(probID_file, make_file_name):
    args = generate_args()
    list = []
    with open(probID_file, "r", encoding="utf-8") as f:
        for line in f:
            label, ID = line.split("_")
            ID = int(ID)
            if label == args.label1:
                wb = openpyxl.load_workbook(args.dataset1)
                ws = wb.worksheets[0]
                probID = ws.cell(ID, 1)
                eng = ws.cell(ID, 2)
                comm = ws.cell(ID, 5)
                position = ws.cell(ID, 4)
                t = "P_" + probID.value + "\t" + eng.value + "\t" +\
                    comm.value + "\t" + position.value + "\n"
                list.append(t)
                wb.close()
            elif label == args.label2:
                wb = openpyxl.load_workbook(args.dataset2)
                ws = wb.worksheets[0]
                probID = ws.cell(ID, 1)
                eng = ws.cell(ID, 2)
                comm = ws.cell(ID, 5)
                position = ws.cell(ID, 4)
                t = "S_" + probID.value + "\t" + eng.value + "\t" + \
                    comm.value + "\t" + position.value + "\n"
                list.append(t)
                wb.close()
            else:
                t = "no such label " + label + " \n"
                list.append(t)
    with open(make_file_name, "w", encoding="utf-8") as f:
        for line in list:
            f.write(line)

if __name__ == "__main__":
    args = sys.argv
    prob_file = args[1]
    make_file = args[2]
    make_problem_for_probID(prob_file, make_file)
